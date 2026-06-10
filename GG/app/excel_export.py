"""
excel_export.py — centralised Excel export helpers for Green Garden.

Each function receives a queryset (already filtered) and returns an
HttpResponse with the .xlsx attachment.
"""
import datetime

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────── shared helpers ───────────────

_DARK   = "2F3640"   # header background (matches sidebar)
_ACCENT = "2D8CF0"   # blue accent
_WHITE  = "FFFFFF"
_GREEN  = "E8F8E8"   # paid / active row tint
_RED    = "FDE8E8"   # unpaid / cancelled row tint
_GREY   = "F0F0F0"   # alternate row

_thin = Side(style="thin", color="CCCCCC")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header_style(cell, text):
    cell.value      = text
    cell.font       = Font(bold=True, color=_WHITE, size=10)
    cell.fill       = PatternFill("solid", fgColor=_DARK)
    cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border     = _border


def _data_style(cell, fill_color=None):
    cell.alignment = Alignment(vertical="center", wrap_text=False)
    cell.border    = _border
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=fill_color)


def _auto_width(ws, min_w=8, max_w=35):
    for col in ws.columns:
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


def _title_row(ws, title, col_count):
    ws.insert_rows(1)
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font      = Font(bold=True, size=13, color=_ACCENT)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)

    date_cell = ws.cell(row=2, column=1,
                        value=f"Exported: {datetime.datetime.now().strftime('%B %d, %Y %I:%M %p')}")
    date_cell.font      = Font(italic=True, size=9, color="888888")
    date_cell.alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16


def _make_response(wb, filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _v(value, fallback="—"):
    """Return value or fallback if falsy."""
    return value if value not in (None, "", 0) else fallback


# ─────────────────────────────────────────────── Lots ─────────────────────────

def export_lots(lots):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lot Records"
    ws.freeze_panes = "A4"   # freeze title + date + header rows

    HEADERS = [
        "#", "Lot Type", "Phase", "Block", "Section",
        "Lot No.", "Column Lvl", "P.A. No.", "Contract No.",
        "Discount", "Eff. Down Pmt", "Interment Date",
        "Date Fully Paid", "P.A. Date",
        "TCT A/R", "Col. Level", "Tomb No.",
        "Lot Owner", "Status",
    ]

    _title_row(ws, "Green Garden — Lot Records", len(HEADERS))

    header_row = 3
    for col, h in enumerate(HEADERS, 1):
        _header_style(ws.cell(row=header_row, column=col), h)
    ws.row_dimensions[header_row].height = 30

    for i, lot in enumerate(lots, 1):
        r = header_row + i

        if lot.is_cancelled:
            tint = _RED
            status = "Cancelled"
        elif not lot.status:
            tint = _GREY
            status = "Completed"
        else:
            tint = _GREEN
            status = "Active"

        row_data = [
            i,
            lot.plan,
            _v(lot.phase),
            _v(lot.block),
            _v(lot.section),
            _v(lot.lot_number) if lot.plan not in ("THS", "THTC") else "—",
            _v(lot.column_level) if lot.plan in ("THS", "THTC") else "N/A",
            _v(lot.pa_number),
            _v(lot.contract_number),
            f"{lot.discount_percent}%" if lot.discount_percent else "—",
            float(lot.effective_down_payment) if lot.effective_down_payment else "—",
            lot.interment_date.strftime("%b %d, %Y") if lot.interment_date else "—",
            lot.date_fully_paid.strftime("%b %d, %Y") if lot.date_fully_paid else "—",
            lot.pa_date.strftime("%b %d, %Y") if lot.pa_date else "—",
            _v(lot.columbarium_type),
            f"Level {lot.columbarium_level}" if lot.columbarium_level else "—",
            _v(lot.tomb_number),
            lot.client.full_name,
            status,
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            _data_style(cell, tint)
            if col == 11 and isinstance(val, float):   # Eff. Down Pmt
                cell.number_format = '#,##0.00'

        ws.row_dimensions[r].height = 18

    _auto_width(ws)
    ws.column_dimensions["A"].width = 5   # # col
    return _make_response(wb, "lot_records.xlsx")


# ─────────────────────────────────────────────── Client Records ───────────────

def export_clients(clients):
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Records"
    ws.freeze_panes = "A4"

    HEADERS = [
        "#", "ID", "First Name", "Middle Name", "Last Name",
        "Contact No.", "Civil Status", "Date of Birth",
        "Religion", "Occupation", "Employer", "Employer Address",
        "Spouse Name", "Spouse DOB", "Spouse Occupation", "Spouse Employer",
        "ID Type", "ID Number", "Date Issued", "Place Issued",
        "Plan", "Plan Status",
    ]

    _title_row(ws, "Green Garden — Client Records", len(HEADERS))

    header_row = 3
    for col, h in enumerate(HEADERS, 1):
        _header_style(ws.cell(row=header_row, column=col), h)
    ws.row_dimensions[header_row].height = 30

    for i, client in enumerate(clients, 1):
        r = header_row + i
        cs = client.clientstatus_set.first()

        if cs:
            plan_name = cs.plan
            if cs.is_cancelled:
                plan_status = "Cancelled"
                tint = _RED
            elif cs.plan == "No Plan":
                plan_status = "Inactive"
                tint = _GREY
            elif cs.status:
                plan_status = "Active"
                tint = _GREEN
            else:
                plan_status = "Completed"
                tint = _GREY
        else:
            plan_name = "No Plan"
            plan_status = "Inactive"
            tint = _GREY

        row_data = [
            i,
            client.pk,
            client.client_first_name,
            _v(client.client_middle_name),
            client.client_last_name,
            client.client_contact_number,
            client.client_civil_status,
            client.client_date_birth.strftime("%b %d, %Y") if client.client_date_birth else "—",
            client.client_religion,
            client.client_occupation,
            client.client_employer_name,
            client.client_employer_address,
            _v(client.client_spouse_name),
            client.client_spouse_date_birth.strftime("%b %d, %Y") if client.client_spouse_date_birth else "—",
            _v(client.client_spouse_occupation),
            _v(client.client_spouse_employer),
            client.client_id_type,
            client.client_id_number,
            client.client_date_issued.strftime("%b %d, %Y") if client.client_date_issued else "—",
            client.client_place_issued,
            plan_name,
            plan_status,
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            _data_style(cell, tint)

        ws.row_dimensions[r].height = 18

    _auto_width(ws)
    ws.column_dimensions["A"].width = 5
    return _make_response(wb, "client_records.xlsx")


# ─────────────────────────────────────────────── Payment History ──────────────

def export_payments(client, client_status, payments):
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"

    summary_rows = [
        ("Client ID",        client.pk),
        ("Full Name",        client.full_name),
        ("Contact",          client.client_contact_number),
        ("Civil Status",     client.client_civil_status),
        ("Date of Birth",    client.client_date_birth.strftime("%b %d, %Y") if client.client_date_birth else "—"),
        ("Address",          client.client_address),
        ("",                 ""),
        ("Plan",             client_status.plan),
        ("Monthly Payment",  float(client_status.monthly_payment)),
        ("Duration",         f"{client_status.duration} months"),
        ("Start Date",       client_status.start_date.strftime("%b %d, %Y") if client_status.start_date else "—"),
        ("Months Remaining", client_status.months_remaining),
        ("Total Paid",       float(client_status.paid_balance)),
        ("Remaining Balance",float(client_status.balance)),
        ("Date Fully Paid",  client_status.date_fully_paid.strftime("%b %d, %Y") if client_status.date_fully_paid else "—"),
        ("Status",           "Cancelled" if client_status.is_cancelled else "Active" if client_status.status else "Completed"),
    ]

    _title_row(ws1, f"Green Garden — Payment Summary: {client.full_name}", 2)

    for i, (label, value) in enumerate(summary_rows, 4):
        label_cell = ws1.cell(row=i, column=1, value=label)
        value_cell = ws1.cell(row=i, column=2, value=value)
        label_cell.font   = Font(bold=True, size=10)
        label_cell.border = _border
        value_cell.border = _border
        if isinstance(value, float):
            value_cell.number_format = '#,##0.00'
            value_cell.alignment = Alignment(horizontal="right")

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 30

    # ── Sheet 2: Payment Schedule ─────────────────────────────────────────
    ws2 = wb.create_sheet("Payment Schedule")
    ws2.freeze_panes = "A4"

    HEADERS = ["#", "Month", "Amount (₱)", "Status", "Date Paid"]
    _title_row(ws2, f"Payment Schedule — {client.full_name}", len(HEADERS))

    header_row = 3
    for col, h in enumerate(HEADERS, 1):
        _header_style(ws2.cell(row=header_row, column=col), h)
    ws2.row_dimensions[header_row].height = 28

    for i, payment in enumerate(payments, 1):
        r = header_row + i
        tint = _GREEN if payment.is_paid else _RED
        row_data = [
            i,
            payment.month.strftime("%B %Y") if payment.month else "—",
            float(payment.amount),
            "Paid" if payment.is_paid else "Unpaid",
            payment.date_paid.strftime("%b %d, %Y %I:%M %p") if payment.date_paid else "—",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=col, value=val)
            _data_style(cell, tint)
            if col == 3:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
        ws2.row_dimensions[r].height = 18

    _auto_width(ws2)
    ws2.column_dimensions["A"].width = 5
    return _make_response(wb, f"payments_{client.pk}_{client.client_last_name}.xlsx")


# ─────────────────────────────────────────────── Bookings ─────────────────────

def export_bookings(bookings):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"
    ws.freeze_panes = "A4"

    HEADERS = [
        "#", "Client Name", "Contact", "Event Type",
        "Booking Date", "Time Slot", "Notes", "Status", "Cancellation Reason",
    ]

    _title_row(ws, "Green Garden — Bookings", len(HEADERS))

    header_row = 3
    for col, h in enumerate(HEADERS, 1):
        _header_style(ws.cell(row=header_row, column=col), h)
    ws.row_dimensions[header_row].height = 28

    for i, b in enumerate(bookings, 1):
        r = header_row + i
        tint = _RED if b.status == "Cancelled" else _GREEN

        row_data = [
            i,
            b.client_name,
            b.contact_number,
            b.event_type,
            b.booking_date.strftime("%b %d, %Y") if b.booking_date else "—",
            b.get_booking_time_display(),
            _v(b.notes),
            b.status,
            _v(b.cancellation_reason),
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            _data_style(cell, tint)

        ws.row_dimensions[r].height = 18

    _auto_width(ws)
    ws.column_dimensions["A"].width = 5
    return _make_response(wb, "bookings.xlsx")
